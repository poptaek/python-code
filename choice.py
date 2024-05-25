import openpyxl
import random

save_path="choice.xlsx"

#기존 엑셀파일 불러오기
wb=openpyxl.load_workbook(save_path, data_only=True)

sheet_data = input("원하는 시트를 입력하세요 (hilling , blessing) : ")

if(sheet_data=='hilling'):


    ws= wb[sheet_data]
    row = list(ws.iter_rows())
    row_count=len(row)
    if row_count > 0:
        random_row_index = random.randint(1, row_count) 
        w=ws.cell(row=random_row_index,column=1)
        print(w.value)
    
    else:
        print("시트에 데이터가 없습니다.")

if(sheet_data=='blessing'):


    ws= wb[sheet_data]
    row = list(ws.iter_rows())
    row_count=len(row)
    if row_count > 0:
        random_row_index = random.randint(1, row_count) 
        w=ws.cell(row=random_row_index,column=1)
        print(w.value)
    
    else:
        print("시트에 데이터가 없습니다.")
